import openpyxl
import os

#To create directory in windows if one doesnt exists
#outPath = "C:\\Users\\syede\\Downloads\\SeleniumOutFiles"
#if not os.path.isdir(outPath):
 #   os.makedirs(outPath)


'''

os.path.dirname(os.path.dirname(__file__))+"\\testCases\\testData\\loginData.xlsx"
inFileName = ".\\testData\\loginData.xlsx"
inSheetName = "Sheet1"

#To merge filename , sheetname with filepath for both input and output files
inPath = "C:\\Users\\syede\\Downloads" # filepath mame
inFileName ="login.xlsx"
inSheetName = "Sheet1"
outPath = "C:\\Users\\syede\\Downloads\\SeleniumOutFiles"
outFileName= "login_output.xlsx"
outSheetName = "output"
inFile = os.path.join(inPath,inFileName)
outFile = os.path.join(outPath, outFileName)
'''

# function to get number of rows from an input file
def getRowCount(inFile,inSheetName):
    workbook = openpyxl.load_workbook(inFile)
    sheet=workbook[inSheetName]
    rows = sheet.max_row
    return rows

# function to get number of columns from an input file
def getColCount(inFile,inSheetName):
    workbook = openpyxl.load_workbook(inFile)
    sheet=workbook[inSheetName]
    cols = sheet.max_column
    return cols

# function to get read the input file based on rownumber and colnumber from an input file
def readXL(inFile, inSheetName, rows, cols):
    workbook = openpyxl.load_workbook(inFile)
    sheet= workbook[inSheetName]
    return sheet.cell(row=rows, column=cols).value

# function to read the input file and write it to an output file based on row and col and then append a value to a specific position
# this is helpful if you are automating a test and want to capture test result for each login or test case result
# see example of Salesforce login for 3 different login credentials and store each test result
def writeXL(inFile, inSheetName, rows, cols, Value):
    workbook = openpyxl.load_workbook(inFile)
    sheet= workbook[inSheetName]
    #sheet.cell(row=rows, column=cols).value
    #workbook.save(outFile)
    sheet.cell(row=rows, column=cols).value= Value
    workbook.save(inFile)

#getRowCount(inFile, inSheetName)
#getColCount(inFile, inSheetName)
#readXL(inFile, inSheetName, 2,3)
#writeXL(inFile, inSheetName, 2,15, "Value")



